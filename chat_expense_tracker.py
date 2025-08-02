import os
import json
import datetime
import re
import pandas as pd
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, ContextTypes, filters
from langchain_core.prompts import ChatPromptTemplate
from langchain.chains import LLMChain
from langchain_openai import ChatOpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from telegram import InputFile
from dotenv import load_dotenv

load_dotenv()

# === LLM Setup ===
llm = ChatOpenAI(
    model="mistralai/mistral-7b-instruct",
    openai_api_key=os.getenv("OPENROUTER_API_KEY"),
    openai_api_base="https://openrouter.ai/api/v1"
)

full_prompt = ChatPromptTemplate.from_template(
    """
    You are a strict JSON extractor and classifier.

    Given this text: '{text}', extract all expenses and numeric amounts, and classify each expense into one of these categories:
    [Groceries, Food, Travel, Transport, Health, Entertainment, Utilities, Rent, Shopping, Sports, Education, Miscellaneous]

    Return ONLY valid JSON in this format:
    [
      {{"description": "...", "amount": ..., "category": "..."}},
      ...
    ]

    Rules:
    - Do NOT include currency symbols.
    - Category must be one from the provided list.
    - Output valid JSON only. No extra text or markdown.
    """
)

category_chain = full_prompt | llm

# === Load categories once ===
with open("categories.json", "r") as f:
    CATEGORY_KEYWORDS = json.load(f)

# === Categorization ===
def categorize_expense(description):
    try:
        result = category_chain.invoke({"description": description})
        if isinstance(result, dict):
            json_text = result.get("text", "")
        else:
            json_text = getattr(result, "content", "")

        match = re.search(r'{.*}', json_text, re.DOTALL)
        if match:
            category_json = json.loads(match.group(0))
            return category_json.get("category", "Other")

    except Exception as e:
        print("❌ Categorization error:", str(e))

    return "Other"

    

#save expense to excel

def save_expense_to_excel(description, amount):
    date = datetime.date.today()
    month = date.strftime("%Y-%m")
    filename = f"expenses_{month}.xlsx"
    category=categorize_expense(description)

    if os.path.exists(filename):
        workbook = load_workbook(filename)
        sheet = workbook.active
    else:
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["Date", "Description", "Amount", "Category"])  # header

    sheet.append([str(date), description, amount, category])
    workbook.save(filename)
 
# Download expense sheet
    
async def download_expense_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    month = datetime.date.today().strftime("%Y-%m")
    filename = f"expenses_{month}.xlsx"
    if os.path.exists(filename):
        with open(filename,"rb") as file:
            await update.message.reply_document(InputFile(file, filename=filename))
    else:
        await update.message.reply_text("No expenses recorded yet this month.")

# === Telegram Handlers ===
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("👋 Hi! Send me your expense in plain text and I’ll categorize and log it for you.")

async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_input = update.message.text
    try:
        response = category_chain.invoke({"text": user_input})
        print("🔎 Raw LLM Response:", response)

        # Get response text
        response_text = response.get("text") if isinstance(response, dict) else getattr(response, "content", "")

        # Extract valid JSON array from response
        json_match = re.search(r'\[.*\]', response_text, re.DOTALL)
        if not json_match:
            raise ValueError("No valid JSON array found in LLM response")

        expenses_json = json_match.group(0)
        expenses = json.loads(expenses_json)

        if not isinstance(expenses, list):
            raise ValueError("Parsed response is not a list")

        confirmation = ""
        for exp in expenses:
            description = exp["description"]
            amount = float(exp["amount"])
            category=exp.get("category","Other")
            save_expense_to_excel(description, amount)
            confirmation += f"✅ {description}: ₹{amount} [{category}]\n"

        await update.message.reply_text("Expenses recorded:\n" + confirmation)

    except Exception as e:
        print("❌ Error:", str(e))
        await update.message.reply_text("❌ I couldn't detect valid expenses. Please try again with a clearer format.")
